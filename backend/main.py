from fastapi import FastAPI, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from groq import Groq
import os
import json
import base64
import tempfile
import io
import re
from dotenv import load_dotenv

load_dotenv()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://nchat-experiment-frontend.onrender.com"],
    allow_credentials= True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = Groq(api_key=os.getenv("GROQ_API_KEY"))

MODEL_ID      = "llama-3.3-70b-versatile"
VISION_MODEL  = "meta-llama/llama-4-scout-17b-16e-instruct"
WHISPER_MODEL = "whisper-large-v3-turbo"

# ── Known diagram types that have templates ────────────────────────
DIAGRAM_KEYWORDS = {
    "india_map":       ["india map","states of india","indian states","india geography"],
    "heart":           ["human heart","heart anatomy","cardiac","heart diagram"],
    "atom":            ["atom","atomic structure","atomic model","bohr model"],
    "dna":             ["dna","dna structure","double helix","nucleotide"],
    "cell":            ["animal cell","plant cell","cell structure","cell biology"],
    "solar_system":    ["solar system","planets","orbit","milky way planets"],
    "water_cycle":     ["water cycle","hydrological cycle","evaporation condensation"],
    "eye":             ["human eye","eye anatomy","eye diagram","eye structure"],
    "ear":             ["human ear","ear anatomy","ear structure","ear diagram"],
    "digestive":       ["digestive system","digestion","stomach intestine"],
}

SYSTEM_PROMPT = """You are Nchat, a precise and helpful AI tutor and assistant.

════════════════════════════════════════════
ACCURACY RULES — absolute priority
════════════════════════════════════════════
- NEVER guess. Always compute step by step before writing any answer.
- For ALL arithmetic: compute manually first, then write the result.
- NEVER skip steps in mathematical derivations.
- NEVER invent formulas, compound names, or reaction products.
- For physics: always include units in every step.
- If uncertain: write "I'm not certain — please verify."

════════════════════════════════════════════
FORMATTING RULES
════════════════════════════════════════════
- Use markdown: **bold**, *italic*, headers, bullets, numbered lists, tables
- ALL math in LaTeX: inline $...$ and block $$...$$
- Chemical formulas always in LaTeX: $H_2O$, $CO_2$, $H_2SO_4$
- Never use LaTeX for non-math content

════════════════════════════════════════════
EXAM PREP MODE
════════════════════════════════════════════
When the user says they are preparing for an exam, studying a topic, or wants
revision/practice on a topic, output in this EXACT structure:

## 📚 Topic: [Topic Name]

### Key Concepts
[2-4 bullet points of the most important ideas]

### Core Formulas
[All relevant formulas in LaTeX, each on its own line with explanation]

### Worked Example
[One fully solved example with every step shown]

### Common Mistakes to Avoid
[2-3 bullet points of common errors]

Then output ALL quiz questions for that topic using QUIZ_DATA format.
Then output ALL relevant WIDGET_SPEC blocks for that topic.

If the user asks about multiple topics, cover BOTH fully,
then output quiz questions covering BOTH topics,
then output a WIDGET_SPEC for EACH topic.

════════════════════════════════════════════
QUIZ FORMAT — CRITICAL RULES
════════════════════════════════════════════
ONLY generate quiz format when user explicitly says:
"give me a quiz", "quiz me", "test me", "practice questions",
"quiz on", "test on", "make a quiz", "create a quiz", "give me questions",
"I am preparing for an exam", "help me prepare", "exam prep", "revision", "study"

NEVER generate quiz format for plain explanation requests.

QUIZ ACCURACY RULES:
1. Compute every arithmetic answer YOURSELF before writing options.
2. Answer letter MUST match the option containing your computed result.
3. All 4 options MUST be different values.
4. Generate exactly 5 questions.
5. Re-verify every Answer: line before finishing.

Output format — NO intro text, NO explanation, NO markdown:
QUIZ_DATA:{"questions":[{"q":"question text","options":{"a":"value","b":"value","c":"value","d":"value"},"answer":"b"}]}

════════════════════════════════════════════
MATH / SCIENCE PROBLEM SOLVING
════════════════════════════════════════════
### Solution
Complete step-by-step solution with LaTeX.

### Answer
Final answer with units.

### Verification
Substitute back to confirm.

### Alternative Method 1: [Different technique name]
COMPLETELY DIFFERENT mathematical approach.

### Alternative Method 2: [Another different technique]

### Key Concept
One sentence.

════════════════════════════════════════════
IMAGE / PHOTO ANALYSIS
════════════════════════════════════════════
### What I See
Transcribe EVERYTHING in the image word for word first.

### Verdict
✅ Correct / ❌ Incorrect / ⚠️ Partially Correct

### Correct Solution
Full solution with LaTeX.

### What Went Wrong
Specific error explanation.

### Alternative Method 1: [Different technique]
### Alternative Method 2: [Another approach]
### Key Formula

════════════════════════════════════════════
INTERACTIVE WIDGETS — DYNAMIC GENERATION
════════════════════════════════════════════
You MUST include WIDGET_SPEC blocks for:
- ANY math formula with variables → calculator widget
- ANY shape/geometry concept → shape widget
- ANY graph/function → graph widget
- Unit conversions → converter widget
- Step-by-step processes → steps widget
- Physics simulations → physics widget
- Comparisons → compare widget

MULTIPLE WIDGETS: Output a WIDGET_SPEC for EACH concept separately.
Place ALL widget specs at the very END of your response, each on its own line.

WIDGET_SPEC:{"type":"calculator","title":"...","fields":[...],"formula":"...","result_unit":"...","formula_display":"..."}
WIDGET_SPEC:{"type":"physics","subtype":"energy_bar","title":"...","params":{...}}
WIDGET_SPEC:{"type":"shape","subtype":"circle","title":"...","fields":[...],"formula":"...","result_unit":"...","formula_display":"..."}

Available types: calculator, graph, converter, steps, table, compare, shape, physics
Shape subtypes: circle, rectangle, triangle, cylinder, sphere
Physics subtypes: energy_bar, projectile, pendulum, wave
"""

DIAGRAM_METADATA_PROMPT = """You are a diagram metadata generator.
Given a diagram type and user question, generate ONLY a JSON object with metadata
for making the diagram interactive.

The JSON must have this exact structure:
{
  "title": "diagram title",
  "description": "1-2 sentence overview",
  "regions": [
    {
      "id": "unique_id",
      "label": "Region Name",
      "color": "#hexcolor",
      "hover_color": "#slightly_darker_hex",
      "description": "2-3 sentence explanation of this region",
      "facts": ["fact 1", "fact 2", "fact 3"],
      "importance": "why this region matters"
    }
  ]
}

Rules:
- Generate metadata for ALL major regions of the diagram
- Colors should be educational and visually distinct
- Descriptions must be accurate and educational
- Facts must be correct and interesting
- Output ONLY the JSON object, no other text
- No markdown, no explanation, just the raw JSON
"""

SVG_GENERATION_PROMPT = """You are an SVG diagram generator for educational content.
Generate a clean, accurate, labeled SVG diagram.

Rules:
- Output ONLY valid SVG code starting with <svg and ending with </svg>
- Use viewBox="0 0 800 600"
- Include clear labels for all parts
- Use educational colors (blues, greens, warm tones)
- Each labeled region must have a data-id attribute matching its label (lowercase, underscores)
- Each region must have a data-label attribute with the display name
- Make it visually clear and suitable for students
- No external fonts, no scripts inside SVG
- Use <g> groups to organize related elements
- Output NOTHING except the SVG code
"""


def detect_diagram_type(message: str) -> str | None:
    msg_lower = message.lower()
    for dtype, keywords in DIAGRAM_KEYWORDS.items():
        for kw in keywords:
            if kw in msg_lower:
                return dtype
    # Check for generic diagram requests
    diagram_words = ["diagram", "anatomy", "structure", "map", "chart", "show me", "visualize", "draw"]
    if any(w in msg_lower for w in diagram_words):
        return "custom"
    return None


def compress_image_b64(b64_data: str, max_size_kb: int = 800) -> str:
    try:
        from PIL import Image
        import io as _io
        image_bytes = base64.b64decode(b64_data)
        if len(image_bytes) / 1024 <= max_size_kb:
            return b64_data
        img = Image.open(_io.BytesIO(image_bytes))
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        max_dim = 1200
        w, h = img.size
        if w > max_dim or h > max_dim:
            ratio = min(max_dim / w, max_dim / h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        quality = 85
        while quality >= 30:
            buffer = _io.BytesIO()
            img.save(buffer, format="JPEG", quality=quality, optimize=True)
            if buffer.tell() / 1024 <= max_size_kb:
                return base64.b64encode(buffer.getvalue()).decode("utf-8")
            quality -= 15
        buffer = _io.BytesIO()
        img.save(buffer, format="JPEG", quality=20, optimize=True)
        return base64.b64encode(buffer.getvalue()).decode("utf-8")
    except Exception:
        return b64_data


def extract_pdf_text(b64_data: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(base64.b64decode(b64_data)))
        pages = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"[Page {i+1}]\n{text.strip()}")
        return "\n\n".join(pages) if pages else "No readable text found."
    except Exception as e:
        return f"Could not read PDF: {e}"


def extract_txt_text(b64_data: str) -> str:
    try:
        return base64.b64decode(b64_data).decode("utf-8", errors="replace")
    except Exception as e:
        return f"Could not read file: {e}"


def extract_excel_text(b64_data: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64_data)), data_only=True)
        result = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cleaned = [str(c) if c is not None else "" for c in row]
                if any(cleaned):
                    rows.append(" | ".join(cleaned))
            result.append(f"## Sheet: {name}\n" + "\n".join(rows))
        return "\n\n".join(result)
    except Exception as e:
        return f"Could not read Excel: {e}"


def validate_quiz_data(quiz_json: dict) -> dict:
    questions = quiz_json.get("questions", [])
    for q in questions:
        question_text = q.get("q", "")
        options = q.get("options", {})
        arith = re.match(
            r"[\w\s]*?(\d+(?:\.\d+)?)\s*([\+\-\×\*\/×÷])\s*(\d+(?:\.\d+)?)",
            question_text
        )
        if arith:
            try:
                a_val = float(arith.group(1))
                op    = arith.group(2)
                b_val = float(arith.group(3))
                if op in ("+",):          correct = a_val + b_val
                elif op in ("-",):        correct = a_val - b_val
                elif op in ("*","×","·"): correct = a_val * b_val
                elif op in ("/","÷"):
                    if b_val != 0:        correct = a_val / b_val
                    else:                 continue
                else:                     continue

                correct_str = str(int(correct)) if correct == int(correct) else str(round(correct, 4))
                for key, val in options.items():
                    val_clean = str(val).strip()
                    try:
                        if float(val_clean) == correct:
                            q["answer"] = key
                            break
                    except ValueError:
                        if val_clean == correct_str:
                            q["answer"] = key
                            break
            except Exception:
                pass
    quiz_json["questions"] = questions
    return quiz_json


def post_process_response(content: str) -> str:
    def fix_quiz_block(match):
        raw = match.group(1)
        try:
            quiz_json = json.loads(raw)
            validated = validate_quiz_data(quiz_json)
            return f"QUIZ_DATA:{json.dumps(validated)}"
        except Exception:
            return match.group(0)
    return re.sub(r"QUIZ_DATA:(\{[\s\S]*?\})(?=\n|$)", fix_quiz_block, content)


def build_messages(user_message, history, image_b64, image_mime,
                   file_b64, file_name, file_mime):
    messages  = [{"role": "system", "content": SYSTEM_PROMPT}]
    use_model = MODEL_ID

    for msg in history[-6:]:
        role = "user" if msg["role"] == "user" else "assistant"
        text = msg.get("content", "").strip()
        if text:
            messages.append({"role": role, "content": text})

    if image_b64:
        compressed  = compress_image_b64(image_b64, max_size_kb=800)
        actual_mime = "image/jpeg" if compressed != image_b64 else image_mime
        messages.append({
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:{actual_mime};base64,{compressed}"}},
                {"type": "text", "text": user_message or (
                    "Analyze this image carefully. First transcribe everything. "
                    "If it contains a math/physics/chemistry solution: give verdict, "
                    "show correct solution step by step, explain error, provide 2 genuinely "
                    "different alternative methods. "
                    "If chemical structure: name, formula, IUPAC name, functional groups. "
                    "If diagram: identify all parts."
                )}
            ]
        })
        use_model = VISION_MODEL
    elif file_b64:
        name_lower = (file_name or "").lower()
        mime_lower = (file_mime or "").lower()
        if "pdf" in mime_lower or name_lower.endswith(".pdf"):
            file_text, file_type = extract_pdf_text(file_b64), "PDF"
        elif "excel" in mime_lower or "spreadsheet" in mime_lower or name_lower.endswith((".xlsx",".xls")):
            file_text, file_type = extract_excel_text(file_b64), "Excel"
        elif "text" in mime_lower or name_lower.endswith(".txt"):
            file_text, file_type = extract_txt_text(file_b64), "text file"
        else:
            file_text, file_type = extract_txt_text(file_b64), "file"
        if len(file_text) > 12000:
            file_text = file_text[:12000] + "\n\n[...truncated...]"
        messages.append({"role": "user", "content": (
            f"User uploaded {file_type} '{file_name}'.\nContent:\n{file_text}\n\n"
            f"Question: {user_message or 'Summarize or analyze.'}"
        )})
    else:
        messages.append({"role": "user", "content": user_message})

    return messages, use_model


def is_ios_request(request: Request) -> bool:
    ua = request.headers.get("user-agent", "").lower()
    return (
        "iphone" in ua or "ipad" in ua or "ipod" in ua
        or request.headers.get("X-No-Stream", "").lower() == "true"
    )


@app.get("/")
def home():
    return {"message": "Nchat running ✅", "model": MODEL_ID}


@app.get("/ping")
def ping():
    return {"status": "awake"}


@app.post("/transcribe")
async def transcribe(audio: UploadFile = File(...)):
    try:
        audio_bytes = await audio.read()
        suffix = ".webm"
        if audio.filename:
            ext = os.path.splitext(audio.filename)[-1]
            if ext: suffix = ext
        if audio.content_type:
            ct = audio.content_type.lower()
            if "mp4"  in ct: suffix = ".mp4"
            elif "mpeg" in ct: suffix = ".mp3"
            elif "ogg"  in ct: suffix = ".ogg"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(audio_bytes)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            tr = client.audio.transcriptions.create(
                model=WHISPER_MODEL, file=(os.path.basename(tmp_path), f),
                response_format="text", language="en",
            )
        os.unlink(tmp_path)
        transcript = tr.strip() if isinstance(tr, str) else tr.text.strip()
        return {"transcript": transcript}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Transcription error: {e}")


@app.post("/diagram")
async def generate_diagram(request: Request):
    """
    Generates an interactive diagram.
    Returns: { type, diagram_type, svg, metadata }
    - For template types: svg = template identifier, metadata = AI-generated JSON
    - For custom types: svg = AI-generated SVG string, metadata = parsed from SVG
    """
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    user_message  = data.get("message", "").strip()
    diagram_type  = data.get("diagram_type", None)

    if not user_message and not diagram_type:
        raise HTTPException(status_code=400, detail="No message or diagram type")

    # Auto-detect if not provided
    if not diagram_type:
        diagram_type = detect_diagram_type(user_message)

    is_template = diagram_type and diagram_type != "custom" and diagram_type in DIAGRAM_KEYWORDS

    if is_template:
        # Generate AI metadata for the template
        meta_messages = [
            {"role": "system", "content": DIAGRAM_METADATA_PROMPT},
            {"role": "user", "content": (
                f"Generate metadata for: {diagram_type.replace('_', ' ').title()}\n"
                f"User question: {user_message}\n"
                f"Make it educational and accurate."
            )}
        ]
        try:
            meta_response = client.chat.completions.create(
                model=MODEL_ID,
                messages=meta_messages,
                max_tokens=2000,
                temperature=0.1,
                stream=False,
            )
            meta_text = meta_response.choices[0].message.content or "{}"
            # Extract JSON from response
            json_match = re.search(r'\{[\s\S]*\}', meta_text)
            metadata = json.loads(json_match.group(0)) if json_match else {}
        except Exception as e:
            metadata = {"title": diagram_type.replace("_"," ").title(), "regions": [], "error": str(e)}

        return JSONResponse(content={
            "type":         "template",
            "diagram_type": diagram_type,
            "svg":          None,
            "metadata":     metadata,
        })

    else:
        # Generate custom SVG using AI
        svg_messages = [
            {"role": "system", "content": SVG_GENERATION_PROMPT},
            {"role": "user", "content": (
                f"Generate an educational SVG diagram for: {user_message}\n"
                f"Make it accurate, labeled, and suitable for students.\n"
                f"Every labeled part must have data-id and data-label attributes."
            )}
        ]
        try:
            svg_response = client.chat.completions.create(
                model=MODEL_ID,
                messages=svg_messages,
                max_tokens=3000,
                temperature=0.1,
                stream=False,
            )
            svg_text = svg_response.choices[0].message.content or ""
            # Extract SVG from response
            svg_match = re.search(r'<svg[\s\S]*?</svg>', svg_text, re.IGNORECASE)
            svg_code  = svg_match.group(0) if svg_match else ""

            # Generate metadata for the custom SVG
            meta_messages = [
                {"role": "system", "content": DIAGRAM_METADATA_PROMPT},
                {"role": "user", "content": (
                    f"Generate metadata for this diagram: {user_message}\n"
                    f"Include educational descriptions for all labeled regions."
                )}
            ]
            meta_response = client.chat.completions.create(
                model=MODEL_ID,
                messages=meta_messages,
                max_tokens=2000,
                temperature=0.1,
                stream=False,
            )
            meta_text   = meta_response.choices[0].message.content or "{}"
            json_match  = re.search(r'\{[\s\S]*\}', meta_text)
            metadata    = json.loads(json_match.group(0)) if json_match else {}

        except Exception as e:
            svg_code = ""
            metadata = {"title": user_message, "regions": [], "error": str(e)}

        return JSONResponse(content={
            "type":         "custom",
            "diagram_type": "custom",
            "svg":          svg_code,
            "metadata":     metadata,
        })


@app.post("/chat")
async def chat(request: Request):
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    user_message = data.get("message",    "").strip()
    history      = data.get("history",    [])
    image_b64    = data.get("image_b64",  None)
    image_mime   = data.get("image_mime", "image/jpeg")
    file_b64     = data.get("file_b64",   None)
    file_name    = data.get("file_name",  "") or ""
    file_mime    = data.get("file_mime",  "") or ""

    if not user_message and not image_b64 and not file_b64:
        raise HTTPException(status_code=400, detail="Empty message")

    messages, use_model = build_messages(
        user_message, history,
        image_b64, image_mime,
        file_b64, file_name, file_mime,
    )

    common_params = dict(
        model=use_model, messages=messages,
        max_tokens=2500, temperature=0.2, top_p=0.85,
    )

    if is_ios_request(request):
        try:
            resp    = client.chat.completions.create(**common_params, stream=False)
            content = resp.choices[0].message.content or ""
            content = post_process_response(content)
            return JSONResponse(
                content={"content": content, "streaming": False},
                headers={"Access-Control-Allow-Origin": "*"},
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    def generate():
        try:
            stream = client.chat.completions.create(**common_params, stream=True)
            full = ""
            for chunk in stream:
                delta = chunk.choices[0].delta.content
                if delta:
                    full += delta
                    yield f"data: {json.dumps({'t': delta})}\n\n"
            corrected = post_process_response(full)
            if corrected != full:
                yield f"data: {json.dumps({'correction': corrected})}\n\n"
            yield "data: [DONE]\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'error': str(e)})}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "X-Accel-Buffering":           "no",
            "Cache-Control":               "no-cache, no-transform",
            "Connection":                  "keep-alive",
            "Content-Type":                "text/event-stream; charset=utf-8",
            "Access-Control-Allow-Origin": "*",
        }
    )